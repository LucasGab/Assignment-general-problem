from pulp import *
from openpyxl import Workbook
from munkres import *
import time
import random
import os
import copy
import gurobipy

#Type a instaces to resolve
folder = 'instances'
os.chdir(f'{folder}/')

excelFile = Workbook()
sheet1 = excelFile.active
sheet1.title = f"{folder}_results"
sheet1.append(("Arquivo","Tamanho Agentes","Tamanho Tarefas","Tempo de resolução","Melhor Solução"))


# The number n of agents and tasks
totalAgent = 0
totalTask = 0

# The agents satisfaction for each task, Ex.: agentsSatisfaction[agent][task]
agentsSatisfaction = []

# The agents cost for each task, Ex.: agentsCost[agent][task]
agentsCost = []

# The agents total capacity, Ex.: agentsCapacity[agent]
agentsCapacity = []

# Getting GUROBI solver (you have to install and have a license)
# https://www.gurobi.com/documentation/7.0/quickstart_linux/software_installation_guid.html
# Download Gurobi on: https://www.gurobi.com/downloads/gurobi-software/
# download the readme, to follow instruction on how to install
# Get a license: https://www.gurobi.com/academia/academic-program-and-licenses/
# Use the license with 'grbgetkey'
# Run /opt/gurobi912/linux64/ sudo python3 setup.py install
# export GUROBI_HOME="/opt/gurobi950/linux64"
# export PATH="${PATH}:${GUROBI_HOME}/bin"
# export LD_LIBRARY_PATH="${LD_LIBRARY_PATH}:${GUROBI_HOME}/lib"
gurobiSolver = getSolver('GUROBI',timeLimit=3*60)
solverName = 'GUROBI'
solverType = "manual"

# Setting the resolve solver
resolveSolver = gurobiSolver

# Returns the satisfaction value of an agent "i" execunting a task "j"
def satisfactionValue(agent,task):
    global agentsSatisfaction
    return agentsSatisfaction[agent][task]

# Returns the cost value of an agent "i" execunting a task "j"
def costValue(agent,task):
    global agentsCost
    return agentsCost[agent][task]

# Returns the total capacity value of an agent "i"
def capacityValue(agent):
    global agentsCapacity
    return agentsCapacity[agent]

# It uses pulp library, to solve with CBC and GUROBI
def resolvePulp(filename):
    global solverName
    global resolveSolver
    global agentsSatisfaction
    global totalAgent
    global totalTask
    global sheet1
    global solverType

    funcOptimization = LpMaximize
    # Defines the problem to resolve
    problem = LpProblem("O_problema_de_designacao_generalizado", funcOptimization)

    # Creates the array of agents and tasks numbers
    # Example for totalAgent = 3:
    # agentsRange = [0,1,2]
    # Example for totalTask = 3:
    # tasksRange = [0,1,2]
    agentsRange = range(totalAgent)
    tasksRange = range(totalTask)

    # Creates the binary variables xij, that defines if an "i" agent execute a "j" task
    # with:
    # xij E {0,1} for i = 1,2,...,n; j = 1,2,...,n
    # In the practice createstotalAgent*totalTask binary variables
    agentsExectution = LpVariable.dicts("agents",(agentsRange,tasksRange),0,1,LpInteger)

    # Defines the satisfaction function that will be using the function optimization 
    # (in this case Maximization defined on line 11).
    # So the funtion is:
    # The summation of Xij*Cij, where
    # Cij is the satisfaction value of "i" agent executing "j" task
    problem += lpSum([satisfactionValue(agent,task) * agentsExectution[agent][task] for agent in agentsRange for task in tasksRange])

    # Defines the restriction for task
    # One task can be executed by just one agent
    # The summation of Xij for i=,1,2,...,n must be equal 1
    for task in tasksRange:
        problem += (lpSum([agentsExectution[agent][task] for agent in agentsRange]) == 1, f"Task_Limit_{task}")

    # Defines the restriction for agent
    # One agent can execute many task until his capacity just one task
    # The summation of Xij*Aij for j=,1,2,...,n must be equal less equal then his total capacity
    for agent in agentsRange:
        problem += (lpSum([(agentsExectution[agent][task]*costValue(agent,task)) for task in tasksRange]) <= capacityValue(agent), f"Agent_Limit_{agent}")

    
    if solverType == "manual":
        # Disable pré proccess
        gurobipy.setParam("Presolve", 0)
        # Disable MIP cuts
        gurobipy.setParam("Cuts", 0)
        # Disable Heuristics
        gurobipy.setParam("Heuristics", 0)

    print(f"\nSolver: {solverName}:\n")
    start = time.time()
    # Solve the problem
    problem.solve(resolveSolver)

    end = time.time()
    total_sum = 0

    sheetResult = excelFile.create_sheet(title=f"Resultados_{filename}_individuais")
    sheetResultIndividual = excelFile.create_sheet(title=f"Resultados_{filename}_individuais_comp")
    sheetResult.append(("Agente","Tarefas"))
    sheetResultIndividual.append(("Agente","Limite Agente","Total de Tarefas","Soma das Tarefas","Capacidade consumida"))
    taskArray = []
    taskSum = 0
    totalTasks = 0
    consumedCapacity = 0
    # Calculates the total sum
    for agent in agentsRange:
        taskArray = []
        taskSum = 0
        totalTasks = 0
        consumedCapacity = 0
        for task in tasksRange:
            if agentsExectution[agent][task].value() == 1:
                # print(f"Agent: {agent} executed the task: {task}")
                taskArray.append(task)
                taskSum += satisfactionValue(agent,task)
                total_sum += satisfactionValue(agent,task)
                consumedCapacity += costValue(agent,task)
                totalTasks += 1
        sheetResult.append([agent] + taskArray)
        sheetResultIndividual.append([agent,capacityValue(agent),totalTasks,taskSum,consumedCapacity])
    
    print(f"The maximization returned a summation of: {total_sum}")
    print(f"Time Elapsed: {end-start}")
    sheet1.append((str(filename),str(totalAgent),str(totalTask),str(end-start),str(total_sum)))

for filename in os.listdir(os.getcwd()):
  excelInstance = Workbook()
  sheetInit = excelInstance.active
  sheetInit.title = f"values"
  sheetInit.append(("Tamanho Agentes","Tamanho Tarefas"))
  with open(os.path.join(os.getcwd(), filename), 'r') as f:
    agentsSatisfaction = []
    agentsCost = []
    agentsCapacity = []
    print(f"Reading file: {filename}")
    lineNumber = 0
    agentNumber = 0
    # For each line read the values following the rule:
    # First Line agents (m) and tasks (n)
    # For each agent (i=0,1,2,...,m): reads satisfaction for doing j tasks (n values)
    # For each agent (i=0,1,2,...,m): reads the cost for doing j task (n values)
    # Last line contains m values representing the total capacity of an agent
    for line in f:
        line = line.replace("\n","")
        line = line.replace("\r","")
        if (lineNumber == 0):
            values = line.split()
            totalAgent = int(values[0])
            totalTask = int(values[1])
            sheetInit.append((totalAgent,totalTask))
            print(f"Total Agents: {totalAgent}")
            print(f"Total Tasks: {totalTask}")
        else:
            values = line.split()
            if (agentNumber < totalAgent):
                satisfaction_numbers = [int(x) for x in values]
                agentsSatisfaction.append(satisfaction_numbers)
            elif (agentNumber < 2*totalAgent):
                cost_numbers = [int(x) for x in values]
                agentsCost.append(cost_numbers)
            else:
                agentsCapacity = [int(x) for x in values]

            agentNumber+=1

        lineNumber+=1

    # Write Agent satisfaction sheet
    agentIndex = 0
    sheetSatisfaction = excelInstance.create_sheet(title="Satisfação")
    sheetSatisfaction.append(["Tarefa (Satisfação):"]+list(range(totalTask)))
    sheetSatisfaction.append(("Agente",))
    for agentSatisfaction in agentsSatisfaction:
        sheetSatisfaction.append([agentIndex] + agentSatisfaction)
        agentIndex+=1
    
    # Write Agent cost sheet
    agentIndex = 0
    sheetCost = excelInstance.create_sheet(title="Custo")
    sheetCost.append(["Tarefa (Custo):"]+list(range(totalTask)))
    sheetCost.append(("Agente",))
    for agentCost in agentsCost:
        sheetCost.append([agentIndex] + agentCost)
        agentIndex+=1
    

    # Write Agent capacity sheet
    agentIndex = 0
    sheetCapacity = excelInstance.create_sheet(title="Capacidade")
    sheetCapacity.append(("Agente","Capacidade"))
    for agentCapacity in agentsCapacity:
        sheetCapacity.append((agentIndex,agentCapacity))
        agentIndex+=1

    excelInstance.save(f"../instances_values/{filename}_values.xlsx")
    resolvePulp(filename)
    
    

excelFile.save(f"../{folder}_{solverName}_{solverType}_results.xlsx")
